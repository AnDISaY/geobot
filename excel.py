from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from datetime import datetime


def add_border(sheet):
    def as_text(value):
        if value is None:
            return ""
        return str(value)
    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells) + 10
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
        for cell in column_cells:
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),)


def initialize_sheet(manager_id_to_name_of_sheet, employees):
    wb = Workbook()
    sheet = wb.active

    sheet['A1'] = 'Ф.И.О.'
    sheet['B1'] = 'Присутствие'
    sheet['C1'] = 'Опоздания'
    sheet['D1'] = 'Отпрашивания'
    sheet['E1'] = 'Дата'
    sheet['F1'] = 'Геолокация'

    date = f'{datetime.now().day}.{datetime.now().month}.{datetime.now().year}'

    for name in employees:
        sheet.cell(row=sheet.max_row+1, column=1, value=name)
        sheet.cell(row=sheet.max_row, column=7, value=date)

    add_border(sheet)

    wb.save(f"sheets/Отчет -- {manager_id_to_name_of_sheet}.xlsx")


def add_attendance(manager_id_to_name_of_sheet, data, employee_name):
    wb = load_workbook(f'sheets/Отчет -- {manager_id_to_name_of_sheet}.xlsx')
    sheet = wb.active
    cell_row = 1
    for row in sheet.rows:
        for cell in row:
            # print(cell.value)
            if cell.value == employee_name:
                cell_row = cell.row

    for k, v in data.items():
        sheet.cell(row=cell_row, column=k, value=v)
    # wb = load_workbook(f'{name_of_sheet}.xlsx')

    add_border(sheet)
    wb.save(f'sheets/Отчет -- {manager_id_to_name_of_sheet}.xlsx')


def update_row(name_of_sheet, time=None):
    pass


# initialize_sheet(772658015, ('bebebe', 'bobobo'))